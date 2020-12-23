#! python3

### Excel
#pip install openpyxl
import openpyxl
import os

os.chdir('c:Python39\\Python_notes')

workbook = openpyxl.load_workbook('example.xlsx')
type(workbook)
sheet = workbook.get_sheet_by_name('sheet1')
type(sheet)
workbook.get_sheet_names() # returns a list will all the sheets of the file
#carefull here
cell  = sheet['A1'] #sheet[index] evaluates to a cell object
cell.value #that cell object has a variable named value, that has the value
sheet['A1'].value #or i can just do that!

sheet.cell(row = 1, column = 2)
#is the same as
sheet['B1']
# .cell is more convenient for "for loops"

#Writing in excel
import openpyxl
wb = openpyxl.workbook()
wb
wb.get_sheet_by_name('Sheet')
sheet['A1'].value
sheet['A1'].value == None

sheet['A1'] = 42
sheet['A2'] = 'Hello'
import os
os.chdir('c:\\Users\\Aristos\\Desktop')
wb.save(example.xlxs) #finaly i have to save it, since thus far is only created in pyhton.

#if i wanted to create more sheets (every wb starts with 3 sheets)
sheet2 = wb.create_sheet()
wb.get_sheet_names()
sheet2.title
sheet2.title = 'New Sheet'
wb.get_sheet_names()
wb.save('example2.xlxs')

wb.create_sheet(index = 0, title = "my other sheet")
wb.save(example3.xlxs)